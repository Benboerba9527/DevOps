import random
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
import os
import sys

def generate_problem():
    while True:
        op = random.choice(['+', '-'])
        a = random.randint(1, 10)
        b = random.randint(1, 10)
        if op == '+':
            answer = a + b
        else:
            answer = a - b
        if 1 <= answer <= 10:
            return f"{a} {op} {b} ="

def main():
    problems = []
    for _ in range(30):
        while True:
            prob = generate_problem()
            if prob not in problems:
                problems.append(prob)
                break

    problems_col1 = problems[:15]
    problems_col2 = problems[15:]

    wb = Workbook()
    ws = wb.active
    ws.title = "MathTest"

    align_left = Alignment(horizontal='left', vertical='center')
    bold_font = Font(bold=True, size=13)

    # 添加大标题
    today_str = datetime.now().strftime("%Y-%m-%d")
    title = f"{today_str} 算数练习题：十以内加减法"
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = bold_font
    ws.row_dimensions[1].height = 28  # 设置标题行高

    # 表头
    ws.append(["ID", "题目", "", "ID", "题目"])

    # 填充内容
    for i in range(15):
        row = [
            i + 1, problems_col1[i], "", i + 16, problems_col2[i]
        ]
        ws.append(row)

    # 设置所有单元格左对齐
    font_24 = Font(size=24)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = align_left
            cell.font = font_24

    # 自动调整列宽
    ws.column_dimensions['A'].width = 4   # ID列宽度固定
    ws.column_dimensions['D'].width = 4   # 右侧ID列宽度固定
    for col in ['B', 'E']:
        max_length = 0
        for cell in ws[col]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        # 24号字体下适当加大宽度
        ws.column_dimensions[col].width = max_length * 2 + 4
    ws.column_dimensions['C'].width = 12  # 空列宽度更大

    # 生成带时间戳的文件名
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"math_test_{now_str}.xlsx"
    wb.save(filename)
    print(f"已生成文件: {filename}")

    # Windows下自动打开文件
    if sys.platform.startswith('win'):
        os.startfile(filename)

if __name__ == "__main__":
    main()